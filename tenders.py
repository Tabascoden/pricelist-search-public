#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
from datetime import datetime, timezone
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Sequence, Tuple

import psycopg2
import psycopg2.extras
from flask import Blueprint, jsonify, render_template, request, send_file

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None

APP_TITLE = os.getenv("APP_TITLE", "iirest")

SCORE_THRESHOLD = float(os.getenv("TENDERS_SCORE_THRESHOLD", "0.20"))
DEFAULT_OFFERS_LIMIT = int(os.getenv("TENDERS_OFFERS_LIMIT", "30"))
MAX_OFFERS_LIMIT = 100
XLSX_HEADER_SCAN_ROWS = 80  # сколько первых строк смотреть в поиске заголовков


def db_connect():
    return psycopg2.connect(
        host=os.getenv("DB_HOST", "127.0.0.1"),
        port=int(os.getenv("DB_PORT", "5432")),
        dbname=os.getenv("DB_NAME", "smartproc"),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASSWORD", ""),
        connect_timeout=5,
        options="-c statement_timeout=12000",
    )


def normalize_query(val: Optional[str]) -> str:
    """Нормализация строки для pg_trgm сравнения (без unaccent)."""
    if val is None:
        return ""
    text = str(val).lower().replace("ё", "е")
    text = re.sub(r"[^0-9a-zа-я\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


tenders_bp = Blueprint("tenders", __name__)  # имя blueprint лучше без _bp


def _project_exists(conn, project_id: int) -> bool:
    with conn.cursor() as cur:
        cur.execute("SELECT 1 FROM tender_projects WHERE id=%s;", (project_id,))
        return cur.fetchone() is not None


def _fetch_projects(conn) -> List[Dict[str, Any]]:
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            """
            SELECT tp.id, tp.title, tp.created_at,
                   COALESCE(items.items_count, 0) AS items_count
            FROM tender_projects tp
            LEFT JOIN LATERAL (
                SELECT count(*) AS items_count FROM tender_items ti WHERE ti.project_id = tp.id
            ) items ON TRUE
            ORDER BY tp.created_at DESC, tp.id DESC;
            """
        )
        return [dict(r) for r in cur.fetchall()]


def _fetch_project(conn, project_id: int) -> Optional[Dict[str, Any]]:
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            "SELECT id, title, created_at FROM tender_projects WHERE id=%s;",
            (project_id,),
        )
        project = cur.fetchone()
        if not project:
            return None

        cur.execute(
            """
            SELECT ti.id, ti.project_id, ti.row_no,
                   ti.name_input, ti.name_raw, ti.qty, ti.unit_input, ti.unit_raw,
                   sel.id AS selected_offer_id,
                   sel.supplier_name, sel.item_name, sel.price, sel.price_per_unit,
                   sel.base_unit, sel.base_qty, sel.score, sel.chosen_at
            FROM tender_items ti
            LEFT JOIN LATERAL (
                SELECT toff.*
                FROM tender_offers toff
                WHERE toff.tender_item_id = ti.id
                ORDER BY toff.chosen_at DESC NULLS LAST, toff.id DESC
                LIMIT 1
            ) sel ON TRUE
            WHERE ti.project_id=%s
            ORDER BY ti.row_no ASC NULLS LAST, ti.id ASC;
            """,
            (project_id,),
        )
        items = [dict(r) for r in cur.fetchall()]

    result = dict(project)
    result["items"] = items
    return result


def _json_safe(v: Any) -> Any:
    # Decimal / datetime безопасно сериализуем
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    if isinstance(v, datetime):
        return v.isoformat()
    try:
        return float(v)
    except Exception:
        return v


def _serialize_project(project: Dict[str, Any]) -> Dict[str, Any]:
    data = {k: _json_safe(v) for k, v in project.items() if k != "items"}
    data["items"] = [{k: _json_safe(v) for k, v in it.items()} for it in project.get("items", [])]
    return data


def _parse_xlsx(stream) -> Tuple[List[Tuple[int, str, float, str]], int]:
    """
    Практичный парсер:
    - ищет строку заголовков, где в ОДНОЙ строке есть "наимен/товар/номенклат" и "кол/количество"
    - дальше читает строки ниже
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl is not available")

    wb = load_workbook(stream, read_only=True, data_only=True)
    ws = wb.active

    name_keywords = ("наимен", "товар", "номенклат", "name", "item", "product")
    qty_keywords = ("кол", "кол-во", "количество", "qty", "quantity")
    unit_keywords = ("ед", "ед.", "единиц", "unit")

    header_row_idx = None
    name_idx = qty_idx = unit_idx = None

    # Ищем заголовок: индексы должны быть найдены В ТОЙ ЖЕ строке
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > XLSX_HEADER_SCAN_ROWS:
            break
        values = [str(v).strip() if v is not None else "" for v in row]
        lowered = [v.lower() for v in values]

        row_name_idx = None
        row_qty_idx = None
        row_unit_idx = None

        for idx, cell_val in enumerate(lowered):
            if row_name_idx is None and any(k in cell_val for k in name_keywords):
                row_name_idx = idx
            if row_qty_idx is None and any(k in cell_val for k in qty_keywords):
                row_qty_idx = idx
            if row_unit_idx is None and any(k in cell_val for k in unit_keywords):
                row_unit_idx = idx

        if row_name_idx is not None and row_qty_idx is not None:
            header_row_idx = row_idx
            name_idx = row_name_idx
            qty_idx = row_qty_idx
            unit_idx = row_unit_idx
            break

    if header_row_idx is None or name_idx is None or qty_idx is None:
        raise ValueError("Не нашёл строку заголовков (нужны колонки Наименование и Количество).")

    start_row = header_row_idx + 1

    rows_to_insert: List[Tuple[int, str, float, str]] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=1):
        vals: Sequence[Any] = list(row)

        # name
        name_val = ""
        if name_idx < len(vals) and vals[name_idx] is not None:
            name_val = str(vals[name_idx]).strip()
        if not name_val:
            continue

        # qty default = 1.0
        qty_val = 1.0
        if qty_idx < len(vals) and vals[qty_idx] is not None and str(vals[qty_idx]).strip():
            try:
                qty_val = float(str(vals[qty_idx]).replace(",", "."))
            except Exception:
                qty_val = 1.0

        # unit
        unit_val = ""
        if unit_idx is not None and unit_idx < len(vals) and vals[unit_idx] is not None:
            unit_val = str(vals[unit_idx]).strip()

        rows_to_insert.append((row_no, name_val, qty_val, unit_val))

    return rows_to_insert, len(rows_to_insert)


def _snapshot_from_supplier_item(
    supplier_item: Dict[str, Any],
    supplier_name: str,
    tender_item: Dict[str, Any],
    score: Optional[float],
) -> Dict[str, Any]:
    return {
        "project_id": tender_item.get("project_id"),
        "tender_item_id": tender_item.get("id"),
        "offer_type": "selected",
        "supplier_id": supplier_item.get("supplier_id"),
        "supplier_item_id": supplier_item.get("id"),
        "supplier_name": supplier_name,
        "item_name": supplier_item.get("name_raw"),
        "name_raw": tender_item.get("name_raw") or tender_item.get("name_input"),
        "unit": supplier_item.get("unit"),
        "price": supplier_item.get("price"),
        "base_unit": supplier_item.get("base_unit"),
        "base_qty": supplier_item.get("base_qty"),
        "price_per_unit": supplier_item.get("price_per_unit"),
        "score": score,
        "category_id": supplier_item.get("category_id"),
        "chosen_at": datetime.now(timezone.utc),
    }


# ---------------- Pages ----------------

@tenders_bp.route("/tenders", methods=["GET"])
def page_tenders():
    # Дадим шаблону проекты (даже если он грузит через JS — лишним не будет)
    try:
        with db_connect() as conn:
            projects = _fetch_projects(conn)
    except Exception:
        projects = []
    return render_template(
        "tenders.html",
        title=f"{APP_TITLE} — Тендеры",
        active="tenders",
        projects=projects,
    )


@tenders_bp.route("/tenders/<int:project_id>", methods=["GET"])
def page_tender_detail(project_id: int):
    # Дадим шаблону project/items для совместимости
    try:
        with db_connect() as conn:
            project = _fetch_project(conn, project_id)
    except Exception:
        project = None

    return render_template(
        "tender_project.html",
        title=f"{APP_TITLE} — Тендер #{project_id}",
        active="tenders",
        project_id=project_id,
        project=project,
        items=(project or {}).get("items", []),
    )


# ---------------- API ----------------

@tenders_bp.route("/api/tenders", methods=["GET"])
def api_tenders_list():
    try:
        with db_connect() as conn:
            rows = _fetch_projects(conn)
        return jsonify({"ok": True, "projects": rows})
    except Exception as e:
        return jsonify({"ok": False, "error": "failed to load projects", "details": str(e)}), 500


@tenders_bp.route("/api/tenders", methods=["POST"])
def api_tenders_create():
    data = request.get_json(silent=True) or {}
    title = (data.get("title") or "").strip() or "Тендер"
    try:
        with db_connect() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "INSERT INTO tender_projects(title) VALUES (%s) RETURNING id, title, created_at;",
                    (title,),
                )
                project = cur.fetchone()
            conn.commit()
        return jsonify({"ok": True, "project": dict(project)})
    except Exception as e:
        return jsonify({"ok": False, "error": "failed to create project", "details": str(e)}), 500


@tenders_bp.route("/api/tenders/<int:project_id>", methods=["GET"])
def api_tender_get(project_id: int):
    try:
        with db_connect() as conn:
            project = _fetch_project(conn, project_id)
        if not project:
            return jsonify({"ok": False, "error": "not found"}), 404
        return jsonify({"ok": True, "project": _serialize_project(project)})
    except Exception as e:
        return jsonify({"ok": False, "error": "failed to load project", "details": str(e)}), 500


@tenders_bp.route("/api/tenders/<int:project_id>/upload", methods=["POST"])
def api_tenders_upload(project_id: int):
    upload = request.files.get("file")
    if not upload:
        return jsonify({"ok": False, "error": "file is required"}), 400

    try:
        rows, count = _parse_xlsx(upload.stream)
    except Exception as e:
        return jsonify({"ok": False, "error": "failed to parse xlsx", "details": str(e)}), 400

    try:
        with db_connect() as conn:
            if not _project_exists(conn, project_id):
                return jsonify({"ok": False, "error": "project not found"}), 404
            with conn.cursor() as cur:
                cur.execute("DELETE FROM tender_items WHERE project_id=%s;", (project_id,))
                if rows:
                    psycopg2.extras.execute_values(
                        cur,
                        """
                        INSERT INTO tender_items
                          (project_id, row_no, name_input, qty, unit_input, name_raw, unit_raw)
                        VALUES %s;
                        """,
                        [
                            (project_id, r[0], r[1], r[2], r[3], r[1], r[3])
                            for r in rows
                        ],
                        page_size=1000,
                    )
            conn.commit()
        return jsonify({"ok": True, "inserted": count})
    except Exception as e:
        return jsonify({"ok": False, "error": "failed to upload", "details": str(e)}), 500


@tenders_bp.route("/api/tenders/<int:project_id>/items/<int:item_id>/offers", methods=["GET"])
def api_tenders_offers(project_id: int, item_id: int):
    limit = request.args.get("limit", type=int) or DEFAULT_OFFERS_LIMIT
    limit = max(5, min(limit, MAX_OFFERS_LIMIT))

    try:
        with db_connect() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT id, project_id, COALESCE(name_input, name_raw) AS name_val
                    FROM tender_items
                    WHERE id=%s;
                    """,
                    (item_id,),
                )
                item = cur.fetchone()
                if not item or item["project_id"] != project_id:
                    return jsonify({"ok": False, "error": "item not found"}), 404

                query_norm = normalize_query(item["name_val"])
                if not query_norm:
                    return jsonify({"ok": True, "offers": []})

                cur.execute(
                    """
                    SELECT si.id AS supplier_item_id, s.name AS supplier_name,
                           COALESCE(si.name_raw, '') AS item_name,
                           si.price, si.price_per_unit,
                           similarity(COALESCE(si.name_normalized, si.name_raw), %s) AS score,
                           si.base_unit, si.base_qty, si.unit, si.category_id
                    FROM supplier_items si
                    JOIN suppliers s ON s.id = si.supplier_id
                    WHERE COALESCE(si.name_normalized, si.name_raw) % %s
                      AND similarity(COALESCE(si.name_normalized, si.name_raw), %s) >= %s
                      AND si.price_per_unit IS NOT NULL
                    ORDER BY score DESC, si.price_per_unit ASC NULLS LAST
                    LIMIT %s;
                    """,
                    (query_norm, query_norm, query_norm, SCORE_THRESHOLD, limit),
                )
                offers = [dict(r) for r in cur.fetchall()]
        return jsonify({"ok": True, "offers": offers})
    except Exception as e:
        return jsonify({"ok": False, "error": "failed to load offers", "details": str(e)}), 500


@tenders_bp.route("/api/tenders/<int:project_id>/items/<int:item_id>/select", methods=["POST"])
def api_tenders_select(project_id: int, item_id: int):
    data = request.get_json(silent=True) or {}
    supplier_item_id = data.get("supplier_item_id")
    if not supplier_item_id:
        return jsonify({"ok": False, "error": "supplier_item_id is required"}), 400

    try:
        supplier_item_id = int(supplier_item_id)
    except Exception:
        return jsonify({"ok": False, "error": "supplier_item_id must be int"}), 400

    try:
        with db_connect() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT * FROM tender_items WHERE id=%s;", (item_id,))
                tender_item = cur.fetchone()
                if not tender_item or tender_item.get("project_id") != project_id:
                    return jsonify({"ok": False, "error": "item not found"}), 404

                query_norm = normalize_query(tender_item.get("name_input") or tender_item.get("name_raw") or "")
                cur.execute(
                    """
                    SELECT si.*, s.name AS supplier_name,
                           similarity(COALESCE(si.name_normalized, si.name_raw), %s) AS score
                    FROM supplier_items si
                    JOIN suppliers s ON s.id = si.supplier_id
                    WHERE si.id=%s;
                    """,
                    (query_norm, supplier_item_id),
                )
                supplier_item = cur.fetchone()
                if not supplier_item:
                    return jsonify({"ok": False, "error": "supplier item not found"}), 404

                snap = _snapshot_from_supplier_item(
                    supplier_item,
                    supplier_item.get("supplier_name") or "",
                    tender_item,
                    supplier_item.get("score"),
                )

                cur.execute(
                    """
                    INSERT INTO tender_offers (
                        project_id, tender_item_id, offer_type,
                        supplier_id, supplier_item_id, supplier_name,
                        item_name, name_raw, unit, price, base_unit, base_qty, price_per_unit, score,
                        category_id, chosen_at
                    )
                    VALUES (
                        %(project_id)s, %(tender_item_id)s, %(offer_type)s,
                        %(supplier_id)s, %(supplier_item_id)s, %(supplier_name)s,
                        %(item_name)s, %(name_raw)s, %(unit)s, %(price)s, %(base_unit)s, %(base_qty)s, %(price_per_unit)s, %(score)s,
                        %(category_id)s, %(chosen_at)s
                    )
                    RETURNING id, supplier_name, item_name, price, price_per_unit, score, chosen_at, supplier_item_id;
                    """,
                    snap,
                )
                selected = cur.fetchone()

                cur.execute(
                    "UPDATE tender_items SET selected_offer_id=%s WHERE id=%s;",
                    (selected["id"], item_id),
                )
            conn.commit()
        return jsonify({"ok": True, "selected": dict(selected)})
    except Exception as e:
        return jsonify({"ok": False, "error": "failed to select offer", "details": str(e)}), 500


@tenders_bp.route("/api/tenders/<int:project_id>/autopick", methods=["POST"])
def api_tenders_autopick(project_id: int):
    selected_count = 0
    items_count = 0

    try:
        with db_connect() as conn:
            if not _project_exists(conn, project_id):
                return jsonify({"ok": False, "error": "project not found"}), 404

            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT * FROM tender_items
                    WHERE project_id=%s
                    ORDER BY row_no ASC NULLS LAST, id ASC;
                    """,
                    (project_id,),
                )
                items = cur.fetchall()

                for item in items:
                    items_count += 1
                    query_norm = normalize_query(item.get("name_input") or item.get("name_raw") or "")
                    if not query_norm:
                        continue

                    cur.execute(
                        """
                        SELECT si.*, s.name AS supplier_name,
                               similarity(COALESCE(si.name_normalized, si.name_raw), %s) AS score
                        FROM supplier_items si
                        JOIN suppliers s ON s.id = si.supplier_id
                        WHERE COALESCE(si.name_normalized, si.name_raw) % %s
                          AND similarity(COALESCE(si.name_normalized, si.name_raw), %s) >= %s
                          AND si.price_per_unit IS NOT NULL
                        ORDER BY score DESC, si.price_per_unit ASC NULLS LAST
                        LIMIT 1;
                        """,
                        (query_norm, query_norm, query_norm, SCORE_THRESHOLD),
                    )
                    best = cur.fetchone()
                    if not best:
                        continue

                    snap = _snapshot_from_supplier_item(
                        best, best.get("supplier_name") or "", item, best.get("score")
                    )
                    cur.execute(
                        """
                        INSERT INTO tender_offers (
                            project_id, tender_item_id, offer_type,
                            supplier_id, supplier_item_id, supplier_name,
                            item_name, name_raw, unit, price, base_unit, base_qty, price_per_unit, score,
                            category_id, chosen_at
                        )
                        VALUES (
                            %(project_id)s, %(tender_item_id)s, %(offer_type)s,
                            %(supplier_id)s, %(supplier_item_id)s, %(supplier_name)s,
                            %(item_name)s, %(name_raw)s, %(unit)s, %(price)s, %(base_unit)s, %(base_qty)s, %(price_per_unit)s, %(score)s,
                            %(category_id)s, %(chosen_at)s
                        )
                        RETURNING id;
                        """,
                        snap,
                    )
                    sel_id = cur.fetchone()["id"]
                    cur.execute(
                        "UPDATE tender_items SET selected_offer_id=%s WHERE id=%s;",
                        (sel_id, item["id"]),
                    )
                    selected_count += 1

            conn.commit()
        return jsonify({"ok": True, "selected": selected_count, "items": items_count})
    except Exception as e:
        return jsonify({"ok": False, "error": "failed to autopick", "details": str(e)}), 500


@tenders_bp.route("/api/tenders/<int:project_id>/export", methods=["GET"])
def api_tenders_export(project_id: int):
    try:
        with db_connect() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT 1 FROM tender_projects WHERE id=%s;", (project_id,))
                if not cur.fetchone():
                    return jsonify({"ok": False, "error": "project not found"}), 404

                cur.execute(
                    """
                    SELECT ti.row_no, COALESCE(ti.name_raw, ti.name_input) AS position_name,
                           ti.qty, COALESCE(ti.unit_raw, ti.unit_input) AS unit,
                           off.supplier_name, off.item_name, off.price_per_unit, off.price, off.score
                    FROM tender_items ti
                    LEFT JOIN LATERAL (
                        SELECT toff.*
                        FROM tender_offers toff
                        WHERE toff.tender_item_id = ti.id
                        ORDER BY toff.chosen_at DESC NULLS LAST, toff.id DESC
                        LIMIT 1
                    ) off ON TRUE
                    WHERE ti.project_id=%s
                    ORDER BY ti.row_no ASC NULLS LAST, ti.id ASC;
                    """,
                    (project_id,),
                )
                rows = [dict(r) for r in cur.fetchall()]
    except Exception as e:
        return jsonify({"ok": False, "error": "failed to export", "details": str(e)}), 500

    header = ["Позиция", "Кол-во", "Ед.", "Поставщик", "Товар", "Цена/ед.", "Сумма", "Score"]

    # XLSX
    if Workbook is not None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tenders"
        ws.append(header)

        for row in rows:
            qty = row.get("qty")
            ppu = row.get("price_per_unit")
            total = None
            try:
                if qty is not None and ppu is not None:
                    total = float(qty) * float(ppu)
            except Exception:
                total = None

            ws.append(
                [
                    row.get("position_name"),
                    qty,
                    row.get("unit"),
                    row.get("supplier_name"),
                    row.get("item_name"),
                    ppu,
                    total,
                    row.get("score"),
                ]
            )

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"tender_{project_id}.xlsx",
        )

    # CSV fallback (правильный текстовый поток)
    out = StringIO()
    import csv
    writer = csv.writer(out)
    writer.writerow(header)

    for row in rows:
        qty = row.get("qty")
        ppu = row.get("price_per_unit")
        total = None
        try:
            if qty is not None and ppu is not None:
                total = float(qty) * float(ppu)
        except Exception:
            total = None

        writer.writerow(
            [
                row.get("position_name"),
                qty,
                row.get("unit"),
                row.get("supplier_name"),
                row.get("item_name"),
                ppu,
                total,
                row.get("score"),
            ]
        )

    data = out.getvalue().encode("utf-8-sig")  # Excel-friendly
    bio = BytesIO(data)
    bio.seek(0)
    return send_file(
        bio,
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"tender_{project_id}.csv",
    )
