#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import hashlib
import json
import math
import os
import re
import shutil
import sys
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

import psycopg2
import psycopg2.extras
from flask import Flask, jsonify, render_template, request, send_file
from werkzeug.utils import secure_filename

import import_price
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except Exception:
    Workbook = None  # fallback на CSV


APP_TITLE = os.getenv("APP_TITLE", "iirest")
MIGRATION_CLI = "--migrate" in sys.argv


def create_app() -> Flask:
    app = Flask(__name__, template_folder="templates")
    app.config["JSON_AS_ASCII"] = False
    app.url_map.strict_slashes = False

    max_mb = int(os.getenv("MAX_UPLOAD_MB", "50"))
    app.config["MAX_CONTENT_LENGTH"] = max_mb * 1024 * 1024

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    UPLOAD_DIR = os.getenv("UPLOAD_DIR", os.path.join(BASE_DIR, "uploads"))
    TENDERS_UPLOAD_DIR = os.path.join(UPLOAD_DIR, "tenders")
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(TENDERS_UPLOAD_DIR, exist_ok=True)

    def _env_bool(name: str, default: bool = False) -> bool:
        val = os.getenv(name)
        if val is None:
            return default
        return val.strip().lower() in {"1", "true", "yes", "on"}

    RUN_MIGRATIONS_ONLY = _env_bool("RUN_MIGRATIONS", False) or MIGRATION_CLI
    AUTO_MIGRATE = _env_bool("AUTO_MIGRATE", False)
    MIGRATIONS_DIR = os.getenv("MIGRATIONS_DIR", os.path.join(BASE_DIR, "db", "migrations"))
    MIGRATION_TABLE = os.getenv("MIGRATION_TABLE", "public.schema_migrations")
    MIGRATION_LOCK_KEY = int(os.getenv("MIGRATION_LOCK_KEY", "424242"))
    MIGRATION_STATEMENT_TIMEOUT_MS = int(os.getenv("MIGRATION_STATEMENT_TIMEOUT_MS", "0"))

    def db_connect():
        return psycopg2.connect(
            host=os.getenv("DB_HOST", "127.0.0.1"),
            port=int(os.getenv("DB_PORT", "5432")),
            dbname=os.getenv("DB_NAME", "smartproc"),
            user=os.getenv("DB_USER", "postgres"),
            password=os.getenv("DB_PASSWORD", ""),
            connect_timeout=5,
            options="-c statement_timeout=12000",
        )

    def _json_safe(v):
        if isinstance(v, Decimal):
            return float(v)
        return v

    def _scalar(row, key=None):
        if row is None:
            return None
        if isinstance(row, dict):
            if key and key in row:
                return row[key]
            if len(row) == 1:
                return next(iter(row.values()))
            raise KeyError(f"Expected key '{key}' in row {row.keys()}")
        return row[0]

    def _safe_remove(path: Optional[str]):
        try:
            if path and os.path.isfile(path) and os.path.exists(path):
                os.remove(path)
        except Exception:
            pass

    def _safe_rmtree(path: Optional[str]):
        try:
            if path and os.path.isdir(path) and os.path.exists(path):
                shutil.rmtree(path, ignore_errors=True)
        except Exception:
            pass

    def ensure_schema_migrations(conn) -> None:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS {MIGRATION_TABLE} (
                  filename   text PRIMARY KEY,
                  applied_at timestamptz NOT NULL DEFAULT now(),
                  checksum   text NULL
                );
                """
            )
            cur.execute(f"ALTER TABLE {MIGRATION_TABLE} ADD COLUMN IF NOT EXISTS checksum text;")
        conn.commit()

    def list_migration_files() -> List[str]:
        migrations_path = Path(MIGRATIONS_DIR)
        if not migrations_path.is_absolute():
            migrations_path = Path(BASE_DIR) / migrations_path
        if not migrations_path.is_dir():
            app.logger.info("Migrations dir not found, skipping auto-migrate")
            return []
        return sorted([path.name for path in migrations_path.glob("*.sql")])

    def get_applied_migrations(conn) -> Dict[str, Optional[str]]:
        with conn.cursor() as cur:
            cur.execute(f"SELECT filename, checksum FROM {MIGRATION_TABLE};")
            return {row[0]: row[1] for row in cur.fetchall()}

    def apply_migration(conn, filename: str, sql_text: str, checksum: str, no_tx: bool) -> None:
        original_autocommit = conn.autocommit
        try:
            if no_tx:
                conn.autocommit = True
                with conn.cursor() as cur:
                    cur.execute(sql_text)
                conn.autocommit = False
            else:
                with conn.cursor() as cur:
                    cur.execute(sql_text)
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    INSERT INTO {MIGRATION_TABLE}(filename, checksum)
                    VALUES (%s, %s)
                    ON CONFLICT (filename) DO NOTHING;
                    """,
                    (filename, checksum),
                )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.autocommit = original_autocommit

    def run_migrations(conn) -> None:
        ensure_schema_migrations(conn)
        timeout_ms = MIGRATION_STATEMENT_TIMEOUT_MS
        if timeout_ms and timeout_ms < 12000:
            timeout_ms = 12000
        with conn.cursor() as cur:
            if timeout_ms > 0:
                cur.execute("SET statement_timeout = %s;", (f"{timeout_ms}ms",))
            else:
                cur.execute("SET statement_timeout = 0;")
        with conn.cursor() as cur:
            cur.execute("SELECT pg_advisory_lock(%s);", (MIGRATION_LOCK_KEY,))
        try:
            applied = get_applied_migrations(conn)
            migrations_path = Path(MIGRATIONS_DIR)
            if not migrations_path.is_absolute():
                migrations_path = Path(BASE_DIR) / migrations_path
            if not migrations_path.is_dir():
                app.logger.info("Migrations dir not found, skipping auto-migrate")
                return
            migration_files = list_migration_files()

            for fname, stored_checksum in applied.items():
                path = migrations_path / fname
                if not path.exists():
                    app.logger.error("Applied migration missing on disk: %s", fname)
                    raise RuntimeError(f"Migration file missing: {fname}")
                sql_text = path.read_text(encoding="utf-8")
                current_checksum = hashlib.sha256(sql_text.encode("utf-8")).hexdigest()
                if stored_checksum is None:
                    with conn.cursor() as cur:
                        cur.execute(
                            f"UPDATE {MIGRATION_TABLE} SET checksum = %s WHERE filename = %s;",
                            (current_checksum, fname),
                        )
                    conn.commit()
                elif stored_checksum != current_checksum:
                    app.logger.error(
                        "Migration checksum mismatch for %s (db=%s, disk=%s)",
                        fname,
                        stored_checksum,
                        current_checksum,
                    )
                    raise RuntimeError(f"Migration checksum mismatch: {fname}")

            pending = [fname for fname in migration_files if fname not in applied]
            app.logger.info("Pending migrations: %s", pending)

            if not pending:
                app.logger.info("No migrations to apply")
                return

            for fname in pending:
                path = migrations_path / fname
                sql_text = path.read_text(encoding="utf-8")
                lines = sql_text.splitlines()
                no_tx = bool(lines and lines[0].strip().upper() == "-- NO_TX")
                checksum = hashlib.sha256(sql_text.encode("utf-8")).hexdigest()
                apply_migration(conn, fname, sql_text, checksum, no_tx)
                app.logger.info("Applied migration %s", fname)
            app.logger.info("Applied %d migrations", len(pending))
        finally:
            with conn.cursor() as cur:
                cur.execute("SELECT pg_advisory_unlock(%s);", (MIGRATION_LOCK_KEY,))
            conn.commit()

    def ensure_schema(conn):
        return

    def ensure_schema_compare(conn):
        ensure_schema(conn)
        return

    def get_category_map(conn) -> Dict[str, int]:
        ensure_schema_compare(conn)
        with conn.cursor() as cur:
            cur.execute("SELECT id, code FROM categories;")
            rows = cur.fetchall()
        return {code: cid for cid, code in rows}

    def normalize_category_value(val: Optional[str]) -> Optional[str]:
        if val is None:
            return None
        v = str(val).strip().lower()
        if not v:
            return None
        mapper = {
            "fresh": "fresh",
            "свеж": "fresh",
            "консерв": "canned",
            "марин": "canned",
            "солен": "canned",
            "вялен": "canned",
            "замороз": "frozen",
            "frozen": "frozen",
        }
        for k, code in mapper.items():
            if k in v:
                return code
        return v if v in ("fresh", "canned", "frozen") else None

    # пробуем подготовить схему на старте
    if AUTO_MIGRATE and not RUN_MIGRATIONS_ONLY:
        try:
            with db_connect() as conn:
                run_migrations(conn)
        except Exception:
            app.logger.exception("Auto-migrate failed")

    # ---------------- Pages ----------------
    @app.route("/", methods=["GET"])
    @app.route("/ui", methods=["GET"])
    def page_search():
        return render_template("search.html", title=APP_TITLE, active="search")

    @app.route("/cart", methods=["GET"])
    def page_cart():
        return render_template("cart.html", title=f"{APP_TITLE} — Корзина", active="cart")

    @app.route("/lists", methods=["GET"])
    def page_lists():
        return render_template("lists.html", title=f"{APP_TITLE} — Поставщики", active="lists")

    @app.route("/tenders", methods=["GET"])
    def page_tenders():
        return render_template("tenders.html", title=f"{APP_TITLE} — Тендеры", active="tenders")

    @app.route("/tenders/<int:project_id>", methods=["GET"])
    def page_tender_detail(project_id: int):
        return render_template(
            "tender_project.html", title=f"{APP_TITLE} — Тендер #{project_id}", active="tenders", project_id=project_id
        )

    @app.route("/favicon.ico", methods=["GET"])
    def favicon():
        return ("", 204)

    # ---------------- Health ----------------
    @app.route("/health", methods=["GET"])
    def health():
        try:
            with db_connect() as conn:
                with conn.cursor() as cur:
                    cur.execute("select 1;")
                    cur.fetchone()
            return jsonify({"status": "ok", "db": "ok"})
        except Exception as e:
            return jsonify({"status": "ok", "db": "error", "details": str(e)}), 200

    # ---------------- API: suppliers ----------------
    @app.route("/api/suppliers", methods=["GET"])
    def api_suppliers():
        try:
            with db_connect() as conn:
                ensure_schema(conn)
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT
                          s.id,
                          s.name,
                          su.last_uploaded_at,
                          su.last_filename,
                          su.last_sheet_mode,
                          su.last_sheets,
                          COALESCE(pl.rows_imported, 0) AS rows_imported
                        FROM suppliers s
                        LEFT JOIN supplier_uploads su ON su.supplier_id = s.id
                        LEFT JOIN LATERAL (
                          SELECT rows_imported
                          FROM price_list_files
                          WHERE supplier_id = s.id
                          ORDER BY id DESC
                          LIMIT 1
                        ) pl ON TRUE
                        ORDER BY s.name NULLS LAST, s.id;
                        """
                    )
                    rows = cur.fetchall()
            suppliers = [{k: _json_safe(v) for k, v in dict(r).items()} for r in rows]
            return jsonify({"suppliers": suppliers})
        except Exception as e:
            app.logger.exception("Failed to list suppliers")
            return jsonify(
                {"error": "failed to list suppliers", "details": str(e)}
            ), 500

    @app.route("/api/suppliers", methods=["POST"])
    def api_create_supplier():
        data = request.get_json(silent=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "name is required"}), 400
        try:
            with db_connect() as conn:
                ensure_schema(conn)
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute("INSERT INTO suppliers(name) VALUES (%s) RETURNING id, name;", (name,))
                    row = cur.fetchone()
                conn.commit()
            return jsonify({"supplier": dict(row)})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    # ---------------- API: tenders ----------------

    def _import_tender_items_from_upload(conn, project_id: int, upload):
        category_map = get_category_map(conn)
        df = pd.read_excel(upload.stream)
        cols = {str(c).strip().lower(): idx for idx, c in enumerate(df.columns)}

        def pick(*names):
            for n in names:
                if n in cols:
                    return cols[n]
            return None

        idx_name = pick("наименование", "name", "товар", "product")
        if idx_name is None:
            raise ValueError("Колонка 'Наименование' не найдена")
        idx_qty = pick("кол-во", "количество", "qty", "quantity")
        idx_unit = pick("ед", "единица", "unit", "ед.")
        idx_cat = pick("категория", "category")

        with conn.cursor() as cur:
            cur.execute("SELECT COALESCE(MAX(row_no), 0) FROM tender_items WHERE project_id=%s;", (project_id,))
            row_no = _scalar(cur.fetchone()) or 0

            items_to_insert = []
            for row in df.itertuples(index=False):
                row_list = list(row)
                name_val = str(row_list[idx_name]).strip() if idx_name is not None else ""
                if not name_val:
                    continue
                qty_val = None
                if idx_qty is not None and idx_qty < len(row_list):
                    try:
                        qty_val = float(row_list[idx_qty]) if row_list[idx_qty] not in (None, "") else None
                    except Exception:
                        qty_val = None
                unit_val = None
                if idx_unit is not None and idx_unit < len(row_list):
                    unit_val = str(row_list[idx_unit]).strip() or None
                cat_val = None
                if idx_cat is not None and idx_cat < len(row_list):
                    cat_val = normalize_category_value(row_list[idx_cat])
                category_id = category_map.get(cat_val) if cat_val else None
                row_no += 1
                items_to_insert.append((project_id, row_no, name_val, qty_val, unit_val, category_id))

            if items_to_insert:
                psycopg2.extras.execute_values(
                    cur,
                    """
                    INSERT INTO tender_items(project_id, row_no, name_input, qty, unit_input, category_id)
                    VALUES %s
                    """,
                    items_to_insert,
                )
        return len(items_to_insert)
    def _calc_offer_totals(offer: Dict[str, Any], tender_qty: Optional[Any]):
        total_price = None
        packs_needed = None

        try:
            qty = float(tender_qty)
        except Exception:
            qty = None

        try:
            base_qty = float(offer.get("base_qty")) if offer.get("base_qty") is not None else None
        except Exception:
            base_qty = None

        try:
            ppu = float(offer.get("price_per_unit")) if offer.get("price_per_unit") is not None else None
        except Exception:
            ppu = None

        try:
            price_val = float(offer.get("price")) if offer.get("price") is not None else None
        except Exception:
            price_val = None

        if qty is not None and base_qty and base_qty > 0:
            packs_needed = math.ceil(qty / base_qty)

        if qty is not None and ppu is not None:
            total_price = ppu * qty
        elif packs_needed is not None and price_val is not None:
            total_price = packs_needed * price_val

        return total_price, packs_needed

    def _load_project(conn, project_id: int):
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT id, title, created_at FROM tender_projects WHERE id=%s;",
                (project_id,),
            )
            project = cur.fetchone()
            if not project:
                return None

            cur.execute(
                """
                SELECT ti.id, ti.project_id, ti.row_no, ti.name_input, ti.search_name, ti.qty, ti.unit_input, ti.category_id, ti.selected_offer_id
                FROM tender_items ti
                WHERE ti.project_id=%s
                ORDER BY ti.row_no ASC, ti.id ASC;
                """,
                (project_id,),
            )
            items = [dict(r) for r in cur.fetchall()]

            cur.execute(
                """
                SELECT toff.*
                FROM tender_offers toff
                WHERE toff.tender_item_id = ANY(%s)
                ORDER BY CASE WHEN toff.offer_type='selected' THEN 0 WHEN toff.offer_type='final' THEN 1 ELSE 2 END,
                         toff.price_per_unit ASC NULLS LAST,
                         toff.id DESC;
                """,
                ([it["id"] for it in items] or [0],),
            )
            offers = [dict(r) for r in cur.fetchall()]
            cur.execute(
                """
                SELECT tps.supplier_id, s.name
                FROM tender_project_suppliers tps
                JOIN suppliers s ON s.id = tps.supplier_id
                WHERE tps.project_id=%s
                ORDER BY s.name;
                """,
                (project_id,),
            )
            suppliers = [dict(r) for r in cur.fetchall()]
        offers_by_item: Dict[int, List[Dict[str, Any]]] = {}
        for off in offers:
            offers_by_item.setdefault(off["tender_item_id"], []).append(off)
        for it in items:
            enriched_offers: List[Dict[str, Any]] = []
            for off in offers_by_item.get(it["id"], []):
                total_price, packs_needed = _calc_offer_totals(off, it.get("qty"))
                enriched = dict(off)
                if total_price is not None:
                    enriched["total_price"] = total_price
                if packs_needed is not None:
                    enriched["packs_needed"] = packs_needed
                enriched["tender_qty"] = it.get("qty")
                enriched_offers.append(enriched)
            it["offers"] = enriched_offers
        project_dict = dict(project)
        project_dict["items"] = items
        project_dict["suppliers"] = suppliers
        return project_dict

    def _get_project_supplier_ids(conn, project_id: int) -> List[int]:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT supplier_id
                FROM tender_project_suppliers
                WHERE project_id=%s
                ORDER BY supplier_id;
                """,
                (project_id,),
            )
            ids = [row[0] for row in cur.fetchall()]
            if ids:
                return ids
            cur.execute("SELECT id FROM suppliers ORDER BY id;")
            return [row[0] for row in cur.fetchall()]

    def _table_columns(conn, table_name: str) -> List[str]:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = 'public' AND table_name=%s
                ORDER BY ordinal_position;
                """,
                (table_name,),
            )
            return [row[0] for row in cur.fetchall()]

    def _rebuild_offers_for_item(
        conn,
        tender_item_id: int,
        project_id: int,
        selected_supplier_item_id: int,
        per_supplier_limit: int = 5,
    ) -> List[int]:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, name_input, category_id
                FROM tender_items
                WHERE id=%s AND project_id=%s;
                """,
                (tender_item_id, project_id),
            )
            item = cur.fetchone()
            if not item:
                return []
            supplier_ids = _get_project_supplier_ids(conn, project_id)

            alt_ids: List[int] = []
            for supplier_id in supplier_ids:
                cur.execute(
                    """
                    SELECT
                      si.id AS supplier_item_id,
                      si.supplier_id,
                      s.name AS supplier_name,
                      si.name_raw,
                      si.unit,
                      si.price,
                      si.base_unit,
                      si.base_qty,
                      si.price_per_unit,
                      si.category_id,
                      similarity(coalesce(si.name_normalized, si.name_raw), %(q)s) AS score
                    FROM supplier_items si
                    JOIN suppliers s ON s.id = si.supplier_id
                    WHERE si.supplier_id=%(supplier_id)s
                      AND si.is_active IS TRUE
                      AND (%(category_id)s IS NULL OR si.category_id = %(category_id)s)
                    ORDER BY similarity(coalesce(si.name_normalized, si.name_raw), %(q)s) DESC,
                             si.price_per_unit ASC NULLS LAST,
                             si.id DESC
                    LIMIT %(limit)s;
                    """,
                    {
                        "supplier_id": supplier_id,
                        "category_id": item.get("category_id"),
                        "q": item.get("name_input"),
                        "limit": per_supplier_limit,
                    },
                )
                for alt in cur.fetchall():
                    if alt["supplier_item_id"] == selected_supplier_item_id:
                        continue
                    snap_alt = {
                        **_snapshot_offer(alt, alt["supplier_name"], item.get("category_id")),
                        "supplier_item_id": alt.get("supplier_item_id"),
                        "score": alt.get("score"),
                    }
                    cur.execute(
                        """
                        SELECT id
                        FROM tender_offers
                        WHERE tender_item_id=%s AND supplier_item_id=%s
                        LIMIT 1;
                        """,
                        (tender_item_id, alt["supplier_item_id"]),
                    )
                    existing_alt = cur.fetchone()
                    if existing_alt:
                        cur.execute(
                            """
                            UPDATE tender_offers
                            SET offer_type='alternative', supplier_id=%(supplier_id)s, supplier_name=%(supplier_name)s,
                                name_raw=%(name_raw)s, unit=%(unit)s, price=%(price)s, base_unit=%(base_unit)s,
                                base_qty=%(base_qty)s, price_per_unit=%(price_per_unit)s, category_id=%(category_id)s,
                                score=%(score)s
                            WHERE id=%(id)s
                            RETURNING id;
                            """,
                            {"id": existing_alt["id"], **snap_alt},
                        )
                    else:
                        cur.execute(
                            """
                            INSERT INTO tender_offers
                              (tender_item_id, offer_type, supplier_id, supplier_item_id, supplier_name, name_raw, unit, price, base_unit, base_qty, price_per_unit, category_id, score)
                            VALUES (%(item_id)s, 'alternative', %(supplier_id)s, %(supplier_item_id)s, %(supplier_name)s, %(name_raw)s, %(unit)s,
                                    %(price)s, %(base_unit)s, %(base_qty)s, %(price_per_unit)s, %(category_id)s, %(score)s)
                            RETURNING id;
                            """,
                            {"item_id": tender_item_id, **snap_alt},
                        )
                    alt_ids.append(_scalar(cur.fetchone(), "id"))

            cur.execute(
                """
                DELETE FROM tender_offers
                WHERE tender_item_id=%s AND offer_type='alternative' AND id <> ALL(%s);
                """,
                (tender_item_id, alt_ids or [0]),
            )
            return alt_ids

    @app.route("/api/tenders", methods=["GET"])
    def api_tenders_list():
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT tp.id, tp.title, tp.created_at,
                          (SELECT count(*) FROM tender_items ti WHERE ti.project_id = tp.id) AS items_count
                        FROM tender_projects tp
                        ORDER BY tp.created_at DESC, tp.id DESC;
                        """
                    )
                    rows = cur.fetchall()
            return jsonify({"projects": [dict(r) for r in rows]})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders", methods=["POST"])
    def api_tenders_create():
        upload = request.files.get("file")
        title = (request.form.get("title") or "Тендер").strip() or "Тендер"
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                if not upload:
                    with conn.cursor() as cur:
                        cur.execute("INSERT INTO tender_projects(title) VALUES (%s) RETURNING id;", (title,))
                        pid = _scalar(cur.fetchone(), "id")
                    conn.commit()
                    with db_connect() as conn2:
                        proj = _load_project(conn2, pid)
                    return jsonify({"project": proj})

                with conn.cursor() as cur:
                    cur.execute("INSERT INTO tender_projects(title) VALUES (%s) RETURNING id;", (title,))
                    pid = _scalar(cur.fetchone(), "id")
                _import_tender_items_from_upload(conn, pid, upload)
                conn.commit()

                with db_connect() as conn2:
                    proj = _load_project(conn2, pid)
                return jsonify({"project": proj})
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        except Exception as e:
            app.logger.exception("Failed to create tender project")
            return jsonify(
                {"error": "failed to create tender project", "details": str(e)}
            ), 500

    @app.route("/api/tenders/<int:project_id>/upload", methods=["POST"])
    def api_tenders_upload(project_id: int):
        upload = request.files.get("file")
        if not upload:
            return jsonify({"error": "file required"}), 400
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor() as cur:
                    cur.execute("SELECT id FROM tender_projects WHERE id=%s;", (project_id,))
                    if not cur.fetchone():
                        return jsonify({"error": "tender project not found"}), 404
                inserted = _import_tender_items_from_upload(conn, project_id, upload)
                conn.commit()
            with db_connect() as conn2:
                proj = _load_project(conn2, project_id)
            return jsonify({"project": proj, "inserted": inserted})
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        except Exception as e:
            app.logger.exception("Failed to upload tender items")
            return jsonify(
                {"error": "failed to upload tender items", "details": str(e)}
            ), 500

    @app.route("/api/tenders/<int:project_id>", methods=["GET"])
    def api_tenders_get(project_id: int):
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                proj = _load_project(conn, project_id)
            if not proj:
                return jsonify({"error": "not found"}), 404
            proj["items"] = [
                {k: _json_safe(v) for k, v in it.items()} if isinstance(it, dict) else it for it in proj.get("items", [])
            ]
            for it in proj.get("items", []):
                if isinstance(it, dict) and "offers" in it:
                    it["offers"] = [{k: _json_safe(v) for k, v in off.items()} for off in it["offers"]]
            return jsonify({"project": proj})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/<int:project_id>/items", methods=["POST"])
    def api_tenders_items_add(project_id: int):
        data = request.get_json(silent=True) or {}
        name_input = str(data.get("name_input") or "").strip()
        if not name_input:
            return jsonify({"error": "name_input required"}), 400

        qty_raw = data.get("qty")
        qty_val = None
        if qty_raw not in (None, ""):
            try:
                qty_val = float(str(qty_raw).replace(",", "."))
            except Exception:
                return jsonify({"error": "qty must be numeric"}), 400

        unit_input = str(data.get("unit_input") or "").strip() or None

        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor() as cur:
                    cur.execute("SELECT id FROM tender_projects WHERE id=%s;", (project_id,))
                    if not cur.fetchone():
                        return jsonify({"error": "tender project not found"}), 404
                    cur.execute(
                        "SELECT COALESCE(MAX(row_no), 0) + 1 FROM tender_items WHERE project_id=%s;",
                        (project_id,),
                    )
                    row_no = _scalar(cur.fetchone())
                    cur.execute(
                        """
                        INSERT INTO tender_items(project_id, row_no, name_input, qty, unit_input)
                        VALUES (%s, %s, %s, %s, %s)
                        RETURNING id;
                        """,
                        (project_id, row_no, name_input, qty_val, unit_input),
                    )
                    item_id = _scalar(cur.fetchone())
                conn.commit()

            return jsonify(
                {
                    "item": {
                        "id": item_id,
                        "project_id": project_id,
                        "row_no": row_no,
                        "name_input": name_input,
                        "qty": qty_val,
                        "unit_input": unit_input,
                    }
                }
            )
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/items/<int:item_id>", methods=["PATCH"])
    def api_tenders_items_update(item_id: int):
        data = request.get_json(silent=True) or {}
        updates = []
        values: List[Any] = []

        if "name_input" in data:
            name_input = str(data.get("name_input") or "").strip()
            if not name_input:
                return jsonify({"error": "name_input required"}), 400
            updates.append("name_input=%s")
            values.append(name_input)

        if "qty" in data:
            qty_raw = data.get("qty")
            if qty_raw in (None, ""):
                qty_val = None
            else:
                try:
                    qty_val = float(str(qty_raw).replace(",", "."))
                except Exception:
                    return jsonify({"error": "qty must be numeric"}), 400
            updates.append("qty=%s")
            values.append(qty_val)

        if "unit_input" in data:
            unit_input = str(data.get("unit_input") or "").strip() or None
            updates.append("unit_input=%s")
            values.append(unit_input)

        if "search_name" in data:
            search_raw = data.get("search_name")
            search_name = str(search_raw or "").strip() or None
            updates.append("search_name=%s")
            values.append(search_name)

        if not updates:
            return jsonify({"error": "no fields to update"}), 400

        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        f"""
                        UPDATE tender_items
                        SET {", ".join(updates)}
                        WHERE id=%s
                        RETURNING id, project_id, row_no, name_input, search_name, qty, unit_input;
                        """,
                        (*values, item_id),
                    )
                    row = cur.fetchone()
                    if not row:
                        return jsonify({"error": "tender item not found"}), 404
                conn.commit()

            return jsonify({"item": dict(row)})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/<int:project_id>", methods=["DELETE"])
    def api_tenders_delete(project_id: int):
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM tender_projects WHERE id=%s RETURNING id;", (project_id,))
                    if not cur.fetchone():
                        return jsonify({"error": "not found"}), 404
                conn.commit()
            return jsonify({"status": "ok"})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/<int:project_id>/suppliers", methods=["GET"])
    def api_tenders_suppliers_get(project_id: int):
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT tps.supplier_id, s.name
                        FROM tender_project_suppliers tps
                        JOIN suppliers s ON s.id = tps.supplier_id
                        WHERE tps.project_id=%s
                        ORDER BY s.name;
                        """,
                        (project_id,),
                    )
                    rows = cur.fetchall()
            suppliers = [dict(r) for r in rows]
            supplier_ids = [r["supplier_id"] for r in suppliers]
            return jsonify({"supplier_ids": supplier_ids, "suppliers": suppliers})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/<int:project_id>/suppliers", methods=["PUT", "POST"])
    def api_tenders_suppliers_put(project_id: int):
        data = request.get_json(silent=True) or {}
        supplier_ids = data.get("supplier_ids") or []
        if not isinstance(supplier_ids, list):
            return jsonify({"error": "supplier_ids must be a list"}), 400
        try:
            supplier_ids = [int(x) for x in supplier_ids if str(x).strip()]
        except Exception:
            return jsonify({"error": "invalid supplier_ids"}), 400
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM tender_project_suppliers WHERE project_id=%s;", (project_id,))
                    if supplier_ids:
                        psycopg2.extras.execute_values(
                            cur,
                            """
                            INSERT INTO tender_project_suppliers(project_id, supplier_id)
                            VALUES %s
                            ON CONFLICT DO NOTHING;
                            """,
                            [(project_id, sid) for sid in supplier_ids],
                        )
                conn.commit()
            return api_tenders_suppliers_get(project_id)
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/<int:project_id>/matrix", methods=["GET"])
    def api_tenders_matrix(project_id: int):
        supplier_ids_raw = (request.args.get("supplier_ids") or "").strip()
        min_score_raw = request.args.get("min_score")
        supplier_ids: List[int] = []
        if supplier_ids_raw:
            try:
                supplier_ids = [int(x) for x in supplier_ids_raw.split(",") if str(x).strip()]
            except Exception:
                return jsonify({"error": "invalid supplier_ids"}), 400
        try:
            min_score = float(min_score_raw) if min_score_raw is not None else 0.0
        except Exception:
            min_score = 0.0
        if not supplier_ids:
            return jsonify({"matrix": {}})
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        WITH suppliers AS (
                          SELECT unnest(%(supplier_ids)s::int[]) AS supplier_id
                        )
                        SELECT
                          ti.id AS tender_item_id,
                          sup.supplier_id,
                          si.id AS supplier_item_id,
                          si.name_raw,
                          si.unit,
                          si.price,
                          si.base_unit,
                          si.base_qty,
                          si.price_per_unit,
                          similarity(coalesce(si.name_normalized, si.name_raw), coalesce(ti.search_name, ti.name_input)) AS score
                        FROM tender_items ti
                        CROSS JOIN suppliers sup
                        LEFT JOIN LATERAL (
                          SELECT *
                          FROM supplier_items si
                          WHERE si.supplier_id = sup.supplier_id
                            AND si.is_active IS TRUE
                            AND (ti.category_id IS NULL OR si.category_id = ti.category_id)
                          ORDER BY similarity(coalesce(si.name_normalized, si.name_raw), coalesce(ti.search_name, ti.name_input)) DESC,
                                   si.price_per_unit ASC NULLS LAST,
                                   si.id DESC
                          LIMIT 1
                        ) si ON TRUE
                        WHERE ti.project_id=%(project_id)s;
                        """,
                        {"supplier_ids": supplier_ids, "project_id": project_id},
                    )
                    rows = cur.fetchall()
            matrix: Dict[str, Dict[str, Dict[str, Any]]] = {}
            for row in rows:
                if row["supplier_item_id"] is None:
                    continue
                score = row.get("score")
                if score is not None and score < min_score:
                    continue
                item_key = str(row["tender_item_id"])
                sup_key = str(row["supplier_id"])
                matrix.setdefault(item_key, {})[sup_key] = {
                    "supplier_id": row["supplier_id"],
                    "supplier_item_id": row["supplier_item_id"],
                    "name_raw": row["name_raw"],
                    "unit": row["unit"],
                    "price": _json_safe(row["price"]),
                    "base_unit": row["base_unit"],
                    "base_qty": _json_safe(row["base_qty"]),
                    "price_per_unit": _json_safe(row["price_per_unit"]),
                    "score": _json_safe(row["score"]),
                }
            return jsonify({"matrix": matrix})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    def _snapshot_offer(row: Dict[str, Any], supplier_name: str, category_id: Optional[int]):
        return {
            "supplier_id": row.get("supplier_id"),
            "supplier_item_id": row.get("id") or row.get("supplier_item_id"),
            "supplier_name": supplier_name,
            "name_raw": row.get("name_raw"),
            "unit": row.get("unit"),
            "price": row.get("price"),
            "base_unit": row.get("base_unit"),
            "base_qty": row.get("base_qty"),
            "price_per_unit": row.get("price_per_unit"),
            "category_id": category_id or row.get("category_id"),
        }

    @app.route("/api/tenders/items/<int:item_id>/select", methods=["POST"])
    def api_tenders_select(item_id: int):
        data = request.get_json(silent=True) or {}
        supplier_item_id = data.get("supplier_item_id")
        project_id = data.get("project_id")
        tender_item_id = data.get("tender_item_id") or item_id
        row_no = data.get("row_no")
        add_to_cart = data.get("add_to_cart")

        if supplier_item_id is None:
            return jsonify({"error": "supplier_item_id is required"}), 400

        try:
            supplier_item_id = int(supplier_item_id)
            project_id = int(project_id) if project_id is not None else None
            tender_item_id = int(tender_item_id) if tender_item_id is not None else None
        except Exception:
            return jsonify({"error": "invalid ids"}), 400

        try:
            row_no = int(row_no) if row_no is not None else None
        except Exception:
            row_no = None
        if add_to_cart is None:
            add_to_cart = True
        if isinstance(add_to_cart, str):
            add_to_cart = add_to_cart.strip().lower() in ("1", "true", "yes", "y")
        else:
            add_to_cart = bool(add_to_cart)
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT id, project_id, row_no, qty, unit_input, category_id, name_input
                        FROM tender_items
                        WHERE id=%s;
                        """,
                        (tender_item_id,),
                    )
                    item = cur.fetchone()
                    if item and project_id is not None and item.get("project_id") != project_id:
                        item = None
                    if not item and row_no is not None and project_id is not None:
                        cur.execute(
                            """
                            SELECT id, project_id, row_no, qty, unit_input, category_id, name_input
                            FROM tender_items
                            WHERE project_id=%s AND row_no=%s
                            ORDER BY id ASC
                            LIMIT 1;
                            """,
                            (project_id, row_no),
                        )
                        item = cur.fetchone()
                    if not item:
                        return jsonify({"error": "tender item not found"}), 404
                    tender_item_id = item["id"]

                    cur.execute(
                        """
                        SELECT si.id, si.supplier_id, s.name AS supplier_name, si.name_raw, si.unit, si.price, si.base_unit, si.base_qty,
                               si.price_per_unit, si.category_id, coalesce(si.name_normalized, si.name_raw) AS norm
                        FROM supplier_items si
                        JOIN suppliers s ON s.id = si.supplier_id
                        WHERE si.id=%s;
                        """,
                        (supplier_item_id,),
                    )
                    base = cur.fetchone()
                    if not base:
                        return jsonify({"error": "supplier item not found"}), 404

                    snap = _snapshot_offer(base, base["supplier_name"], item.get("category_id"))
                    cur.execute(
                        "SELECT similarity(%s, %s) AS score;",
                        (base.get("norm"), item.get("name_input")),
                    )
                    snap["score"] = _scalar(cur.fetchone(), "score")

                    cur.execute(
                        "SELECT id FROM tender_offers WHERE tender_item_id=%s AND supplier_item_id=%s LIMIT 1;",
                        (tender_item_id, supplier_item_id),
                    )
                    existing_selected = cur.fetchone()

                    if existing_selected:
                        cur.execute(
                            """
                            UPDATE tender_offers
                            SET offer_type='selected', supplier_id=%(supplier_id)s, supplier_name=%(supplier_name)s,
                                name_raw=%(name_raw)s, unit=%(unit)s, price=%(price)s, base_unit=%(base_unit)s,
                                base_qty=%(base_qty)s, price_per_unit=%(price_per_unit)s, category_id=%(category_id)s,
                                score=%(score)s
                            WHERE id=%(id)s
                            RETURNING id;
                            """,
                            {"id": existing_selected["id"], **snap},
                        )
                    else:
                        cur.execute(
                            """
                            INSERT INTO tender_offers
                              (tender_item_id, offer_type, supplier_id, supplier_item_id, supplier_name, name_raw, unit, price, base_unit, base_qty, price_per_unit, category_id, score)
                            VALUES (%(item_id)s, 'selected', %(supplier_id)s, %(supplier_item_id)s, %(supplier_name)s, %(name_raw)s, %(unit)s,
                                    %(price)s, %(base_unit)s, %(base_qty)s, %(price_per_unit)s, %(category_id)s, %(score)s)
                            RETURNING id;
                            """,
                            {"item_id": tender_item_id, **snap},
                        )
                    selected_id = _scalar(cur.fetchone(), "id")
                    _rebuild_offers_for_item(conn, tender_item_id, item["project_id"], supplier_item_id)

                    if add_to_cart:
                        cur.execute(
                            "UPDATE tender_items SET selected_offer_id=%s WHERE id=%s;",
                            (selected_id, tender_item_id),
                        )
                conn.commit()
            return jsonify(
                {
                    "ok": True,
                    "selected_offer_id": selected_id if add_to_cart else None,
                    "tender_item_id": tender_item_id,
                    "add_to_cart": add_to_cart,
                }
            )
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/items/<int:item_id>/clear", methods=["POST"])
    def api_tenders_clear(item_id: int):
        data = request.get_json(silent=True) or {}
        project_id = data.get("project_id")
        try:
            project_id = int(project_id) if project_id is not None else None
        except Exception:
            project_id = None
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor() as cur:
                    if project_id is not None:
                        cur.execute(
                            "SELECT project_id FROM tender_items WHERE id=%s;",
                            (item_id,),
                        )
                        row = cur.fetchone()
                        if not row or row[0] != project_id:
                            return jsonify({"error": "not found"}), 404
                    cur.execute(
                        """
                        UPDATE tender_offers
                        SET offer_type='alternative'
                        WHERE tender_item_id=%s AND offer_type IN ('selected', 'final');
                        """,
                        (item_id,),
                    )
                    cur.execute(
                        "UPDATE tender_items SET selected_offer_id=NULL WHERE id=%s;",
                        (item_id,),
                    )
                conn.commit()
            return jsonify({"status": "ok"})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/items/<int:item_id>/offers", methods=["GET"])
    def api_tenders_offers(item_id: int):
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                cur.execute(
                    """
                    SELECT ti.project_id, ti.id, ti.name_input, ti.qty FROM tender_items ti WHERE ti.id=%s;
                    """,
                    (item_id,),
                )
                item = cur.fetchone()
                if not item:
                    return jsonify({"error": "not found"}), 404
                cur.execute(
                    """
                    SELECT * FROM tender_offers
                    WHERE tender_item_id=%s
                    ORDER BY CASE
                               WHEN offer_type='selected' THEN 0
                               WHEN offer_type='final' THEN 1
                               ELSE 2
                             END,
                             CASE WHEN offer_type='alternative' THEN score END DESC NULLS LAST,
                             price_per_unit ASC NULLS LAST,
                             id DESC;
                    """,
                    (item_id,),
                )
                offers = []
                for r in cur.fetchall():
                    offer = dict(r)
                    tender_qty = item.get("qty")
                    total_price, packs_needed = _calc_offer_totals(offer, tender_qty)
                    offer["tender_qty"] = tender_qty
                    if total_price is not None:
                        offer["total_price"] = total_price
                    if packs_needed is not None:
                        offer["packs_needed"] = packs_needed
                    offers.append(offer)
            return jsonify({"item": dict(item), "offers": offers})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/items/<int:item_id>/matches", methods=["GET"])
    def api_tenders_matches(item_id: int):
        supplier_id = request.args.get("supplier_id")
        limit = request.args.get("limit") or "25"
        q_input = (request.args.get("q") or "").strip()
        try:
            supplier_id = int(supplier_id) if supplier_id is not None else None
        except Exception:
            supplier_id = None
        try:
            limit_i = int(limit)
        except Exception:
            limit_i = 25
        limit_i = max(1, min(limit_i, 50))
        if supplier_id is None:
            return jsonify({"error": "supplier_id is required"}), 400
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT id, name_input, search_name, category_id
                        FROM tender_items
                        WHERE id=%s;
                        """,
                        (item_id,),
                    )
                    item = cur.fetchone()
                    if not item:
                        return jsonify({"error": "not found"}), 404
                    q_filter = q_input or None
                    search_name = item.get("search_name") or ""
                    q_similarity = q_filter or (search_name or item.get("name_input") or "")
                    q_like = f"%{q_filter}%" if q_filter else None
                    cur.execute(
                        """
                        SELECT
                          si.id AS supplier_item_id,
                          si.supplier_id,
                          si.name_raw,
                          si.unit,
                          si.price,
                          si.base_unit,
                          si.base_qty,
                          si.price_per_unit,
                          similarity(coalesce(si.name_normalized, si.name_raw), %(q)s) AS score
                        FROM supplier_items si
                        WHERE si.supplier_id=%(supplier_id)s
                          AND si.is_active IS TRUE
                          AND (%(category_id)s IS NULL OR si.category_id = %(category_id)s)
                          AND (%(q_filter)s IS NULL OR coalesce(si.name_normalized, si.name_raw) ILIKE %(q_like)s)
                        ORDER BY similarity(coalesce(si.name_normalized, si.name_raw), %(q)s) DESC,
                                 si.price_per_unit ASC NULLS LAST,
                                 si.id DESC
                        LIMIT %(limit)s;
                        """,
                        {
                            "supplier_id": supplier_id,
                            "category_id": item.get("category_id"),
                            "q": q_similarity,
                            "q_filter": q_filter,
                            "q_like": q_like,
                            "limit": limit_i,
                        },
                    )
                    matches = [dict(r) for r in cur.fetchall()]
            return jsonify({"matches": [{k: _json_safe(v) for k, v in m.items()} for m in matches]})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/items/<int:item_id>/finalize", methods=["POST"])
    def api_tenders_finalize(item_id: int):
        data = request.get_json(silent=True) or {}
        offer_id = data.get("offer_id")
        if not offer_id:
            return jsonify({"error": "offer_id is required"}), 400
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT project_id FROM tender_items WHERE id=%s;",
                        (item_id,),
                    )
                    row = cur.fetchone()
                    if not row:
                        return jsonify({"error": "not found"}), 404
                    cur.execute(
                        "UPDATE tender_offers SET offer_type='final' WHERE id=%s AND tender_item_id=%s;",
                        (offer_id, item_id),
                    )
                    cur.execute(
                        "UPDATE tender_items SET selected_offer_id=%s WHERE id=%s;",
                        (offer_id, item_id),
                    )
                conn.commit()
            return jsonify({"status": "ok"})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/<int:project_id>/export", methods=["POST"])
    def api_tenders_export(project_id: int):
        try:
            payload = request.get_json(silent=True) or {}
            overrides_raw = payload.get("order_qty_overrides") or {}
            order_qty_overrides: Dict[int, float] = {}
            for key, value in overrides_raw.items():
                try:
                    item_id = int(key)
                except Exception:
                    continue
                try:
                    if isinstance(value, str):
                        parsed = float(value.replace(",", "."))
                    else:
                        parsed = float(value)
                except Exception:
                    continue
                if math.isfinite(parsed):
                    order_qty_overrides[item_id] = parsed

            with db_connect() as conn:
                ensure_schema_compare(conn)
                proj = _load_project(conn, project_id)
                if not proj:
                    return jsonify({"error": "not found"}), 404
                cart_rows: List[Dict[str, Any]] = []
                for it in proj.get("items", []):
                    selected_id = it.get("selected_offer_id")
                    if not selected_id:
                        continue
                    offer = None
                    for off in it.get("offers", []):
                        if off.get("id") == selected_id:
                            offer = off
                            break
                    if not offer:
                        continue
                    item_id = it.get("id")
                    tender_qty = it.get("qty")
                    order_qty = order_qty_overrides.get(item_id, tender_qty)
                    supplier_price = offer.get("price")
                    total_val = None
                    try:
                        if order_qty is not None and supplier_price is not None:
                            total_val = float(order_qty) * float(supplier_price)
                    except Exception:
                        total_val = None
                    if total_val is None:
                        total_val = offer.get("total_price")
                    if total_val is None:
                        try:
                            total_val, _ = _calc_offer_totals(offer, order_qty)
                        except Exception:
                            total_val = None
                    cart_rows.append(
                        {
                            "row_no": it.get("row_no"),
                            "name_input": it.get("name_input"),
                            "qty": tender_qty,
                            "unit_input": it.get("unit_input"),
                            "supplier_id": offer.get("supplier_id"),
                            "supplier_name": offer.get("supplier_name"),
                            "name_raw": offer.get("name_raw"),
                            "order_qty": order_qty,
                            "supplier_price": supplier_price,
                            "total_price": total_val,
                        }
                    )

                headers = [
                    "№",
                    "ПОЗИЦИЯ",
                    "КОЛИЧЕСТВО",
                    "ЕД.",
                    "ПОСТАВЩИК",
                    "ТОВАР У ПОСТАВЩИКА",
                    "КОЛИЧЕСТВО ДЛЯ ЗАКАЗА",
                    "ЦЕНА ПОСТАВЩИКА",
                    "СУММА",
                ]

                if Workbook is None:
                    out = BytesIO()
                    writer = csv.writer(out)
                    writer.writerow(headers)
                    for r in cart_rows:
                        writer.writerow(
                            [
                                r.get("row_no"),
                                r.get("name_input"),
                                r.get("qty"),
                                r.get("unit_input"),
                                r.get("supplier_name"),
                                r.get("name_raw"),
                                r.get("order_qty"),
                                r.get("supplier_price"),
                                r.get("total_price"),
                            ]
                        )
                    writer.writerow([])
                    writer.writerow(["ИТОГО ПО ПОСТАВЩИКУ", "ПОЗИЦИЙ", "СУММА"])
                    totals_map: Dict[Any, Dict[str, Any]] = {}
                    for r in cart_rows:
                        sid = r.get("supplier_id")
                        if sid is None:
                            continue
                        row_total = r.get("total_price") or 0
                        entry = totals_map.setdefault(
                            sid,
                            {"supplier_name": r.get("supplier_name"), "items": 0, "total": 0.0},
                        )
                        entry["items"] += 1
                        try:
                            entry["total"] += float(row_total)
                        except Exception:
                            entry["total"] += 0.0
                    totals_list = sorted(totals_map.values(), key=lambda x: x.get("total", 0), reverse=True)
                    grand_total = sum([x.get("total") or 0 for x in totals_list])
                    for t in totals_list:
                        writer.writerow([t.get("supplier_name"), t.get("items"), t.get("total")])
                    writer.writerow(["ИТОГО", len(cart_rows), grand_total])
                    out.seek(0)
                    return send_file(out, mimetype="text/csv", as_attachment=True, download_name="tender.csv")

                wb = Workbook()
                ws = wb.active
                ws.title = "Тендер"
                ws.append(headers)

                header_fill = PatternFill("solid", fgColor="F8FAFC")
                header_font = Font(bold=True)
                border_side = Side(style="thin", color="E2E8F0")
                border = Border(top=border_side, left=border_side, right=border_side, bottom=border_side)

                for c in ws[1]:
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = Alignment(vertical="center", wrap_text=True)
                    c.border = border

                for r in cart_rows:
                    ws.append(
                        [
                            r.get("row_no"),
                            r.get("name_input"),
                            r.get("qty"),
                            r.get("unit_input"),
                            r.get("supplier_name"),
                            r.get("name_raw"),
                            r.get("order_qty"),
                            r.get("supplier_price"),
                            r.get("total_price"),
                        ]
                    )

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if row[1].value is not None:
                        row[1].font = Font(bold=True)
                    if row[8].value is not None:
                        row[8].font = Font(bold=True)

                number_columns = {3: "0.###", 7: "0.###", 8: "#,##0.00", 9: "#,##0.00"}
                for col_idx, fmt in number_columns.items():
                    for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        for c in cell:
                            if isinstance(c.value, (int, float)):
                                c.number_format = fmt

                column_widths = [6, 30, 12, 8, 22, 45, 20, 18, 16]
                for idx, width in enumerate(column_widths, start=1):
                    ws.column_dimensions[chr(64 + idx)].width = width

                ws.append([])
                ws.append(["ИТОГО ПО ПОСТАВЩИКУ", "ПОЗИЦИЙ", "СУММА"])
                totals_header_row = ws.max_row
                for c in ws[totals_header_row]:
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = Alignment(vertical="center")
                    c.border = border

                totals_map: Dict[Any, Dict[str, Any]] = {}
                for r in cart_rows:
                    sid = r.get("supplier_id")
                    if sid is None:
                        continue
                    row_total = r.get("total_price") or 0
                    entry = totals_map.setdefault(
                        sid,
                        {"supplier_name": r.get("supplier_name"), "items": 0, "total": 0.0},
                    )
                    entry["items"] += 1
                    try:
                        entry["total"] += float(row_total)
                    except Exception:
                        entry["total"] += 0.0

                totals_list = sorted(totals_map.values(), key=lambda x: x.get("total", 0), reverse=True)
                grand_total = sum([x.get("total") or 0 for x in totals_list])
                for t in totals_list:
                    ws.append([t.get("supplier_name"), t.get("items"), t.get("total")])
                ws.append(["ИТОГО", len(cart_rows), grand_total])

                totals_start_row = totals_header_row
                totals_end_row = ws.max_row
                for row in ws.iter_rows(min_row=totals_start_row, max_row=totals_end_row, min_col=1, max_col=3):
                    for cell in row:
                        cell.border = border
                        if cell.row != totals_header_row:
                            cell.alignment = Alignment(vertical="top")
                            if cell.col_idx == 3 and isinstance(cell.value, (int, float)):
                                cell.number_format = "#,##0.00"
                    if row[0].value == "ИТОГО":
                        for cell in row:
                            cell.font = Font(bold=True)
                    elif row[0].value is not None and row[0].row != totals_header_row:
                        row[0].font = Font(bold=True)
                        if row[2].value is not None:
                            row[2].font = Font(bold=True)

                bio = BytesIO()
                wb.save(bio)
                bio.seek(0)
                return send_file(
                    bio,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True,
                    download_name=f"tender_{project_id}.xlsx",
                )
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/tenders/<int:project_id>/orders", methods=["POST"])
    def api_tenders_orders(project_id: int):
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT ti.id AS tender_item_id,
                               ti.qty,
                               toff.*
                        FROM tender_items ti
                        LEFT JOIN LATERAL (
                          SELECT *
                          FROM tender_offers toff
                          WHERE toff.tender_item_id = ti.id
                            AND (toff.offer_type = 'final' OR toff.id = ti.selected_offer_id)
                          ORDER BY CASE WHEN toff.offer_type='final' THEN 0 ELSE 1 END
                          LIMIT 1
                        ) toff ON TRUE
                        WHERE ti.project_id=%s;
                        """,
                        (project_id,),
                    )
                    rows = cur.fetchall()

                    by_supplier: Dict[int, List[Dict[str, Any]]] = {}
                    for row in rows:
                        if not row.get("supplier_id") or not row.get("supplier_item_id"):
                            continue
                        by_supplier.setdefault(row["supplier_id"], []).append(row)

                    if not by_supplier:
                        return jsonify({"orders": []})

                    order_columns = set(_table_columns(conn, "orders"))
                    item_columns = set(_table_columns(conn, "order_items"))

                    orders_out: List[Dict[str, Any]] = []
                    for supplier_id, items in by_supplier.items():
                        order_fields = {}
                        if "supplier_id" in order_columns:
                            order_fields["supplier_id"] = supplier_id
                        if "tender_project_id" in order_columns:
                            order_fields["tender_project_id"] = project_id

                        if not order_fields:
                            cur.execute("INSERT INTO orders DEFAULT VALUES RETURNING id;")
                            order_id = _scalar(cur.fetchone(), "id")
                        else:
                            cols = ", ".join(order_fields.keys())
                            placeholders = ", ".join(["%s"] * len(order_fields))
                            cur.execute(
                                f"INSERT INTO orders ({cols}) VALUES ({placeholders}) RETURNING id;",
                                list(order_fields.values()),
                            )
                            order_id = _scalar(cur.fetchone(), "id")

                        total_price = 0.0
                        item_rows = []
                        for it in items:
                            tender_qty = it.get("qty")
                            total_price_item, _ = _calc_offer_totals(it, tender_qty)
                            total_price_item = float(total_price_item) if total_price_item is not None else 0.0
                            total_price += total_price_item
                            row = {}
                            if "order_id" in item_columns:
                                row["order_id"] = order_id
                            if "supplier_item_id" in item_columns:
                                row["supplier_item_id"] = it.get("supplier_item_id")
                            if "tender_item_id" in item_columns:
                                row["tender_item_id"] = it.get("tender_item_id")
                            if "qty" in item_columns:
                                row["qty"] = tender_qty
                            if "price" in item_columns:
                                row["price"] = it.get("price_per_unit") or it.get("price")
                            if "total_price" in item_columns:
                                row["total_price"] = total_price_item
                            if "name_raw" in item_columns:
                                row["name_raw"] = it.get("name_raw")
                            if "unit" in item_columns:
                                row["unit"] = it.get("unit")
                            if row:
                                item_rows.append(row)

                        if item_rows:
                            cols = list(item_rows[0].keys())
                            psycopg2.extras.execute_values(
                                cur,
                                f"INSERT INTO order_items ({', '.join(cols)}) VALUES %s;",
                                [[r.get(c) for c in cols] for r in item_rows],
                            )

                        orders_out.append(
                            {
                                "order_id": order_id,
                                "supplier_id": supplier_id,
                                "items_count": len(items),
                                "total_price": total_price,
                            }
                        )

                    if orders_out:
                        supplier_ids = [o["supplier_id"] for o in orders_out]
                        cur.execute(
                            "SELECT id, name FROM suppliers WHERE id = ANY(%s);",
                            (supplier_ids,),
                        )
                        sup_map = {row["id"]: row["name"] for row in cur.fetchall()}
                        for o in orders_out:
                            o["supplier_name"] = sup_map.get(o["supplier_id"])

                conn.commit()
            return jsonify({"orders": orders_out})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/orders", methods=["GET"])
    def api_orders_list():
        project_id = request.args.get("project_id")
        try:
            project_id = int(project_id) if project_id is not None else None
        except Exception:
            project_id = None
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    if project_id is not None:
                        cur.execute(
                            """
                            SELECT o.*, s.name AS supplier_name
                            FROM orders o
                            LEFT JOIN suppliers s ON s.id = o.supplier_id
                            WHERE o.tender_project_id=%s
                            ORDER BY o.id DESC;
                            """,
                            (project_id,),
                        )
                    else:
                        cur.execute(
                            """
                            SELECT o.*, s.name AS supplier_name
                            FROM orders o
                            LEFT JOIN suppliers s ON s.id = o.supplier_id
                            ORDER BY o.id DESC;
                            """
                        )
                    rows = cur.fetchall()
            orders = [{k: _json_safe(v) for k, v in dict(r).items()} for r in rows]
            return jsonify({"orders": orders})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/orders/<int:order_id>", methods=["GET"])
    def api_orders_get(order_id: int):
        try:
            with db_connect() as conn:
                ensure_schema_compare(conn)
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT o.*, s.name AS supplier_name
                        FROM orders o
                        LEFT JOIN suppliers s ON s.id = o.supplier_id
                        WHERE o.id=%s;
                        """,
                        (order_id,),
                    )
                    order = cur.fetchone()
                    if not order:
                        return jsonify({"error": "not found"}), 404
                    cur.execute(
                        """
                        SELECT oi.*, si.name_raw, si.unit
                        FROM order_items oi
                        LEFT JOIN supplier_items si ON si.id = oi.supplier_item_id
                        WHERE oi.order_id=%s
                        ORDER BY oi.id ASC;
                        """,
                        (order_id,),
                    )
                    items = [dict(r) for r in cur.fetchall()]
            order_out = {k: _json_safe(v) for k, v in dict(order).items()}
            order_out["items"] = [{k: _json_safe(v) for k, v in r.items()} for r in items]
            return jsonify({"order": order_out})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    @app.route("/api/suppliers/<int:supplier_id>", methods=["DELETE"])
    def api_delete_supplier(supplier_id: int):
        stored_path = None
        sup_dir = os.path.join(UPLOAD_DIR, str(supplier_id))
        try:
            with db_connect() as conn:
                ensure_schema(conn)

                with conn.cursor() as cur:
                    cur.execute("SELECT 1 FROM suppliers WHERE id=%s;", (supplier_id,))
                    if not cur.fetchone():
                        return jsonify({"error": "supplier not found"}), 404

                # запрет на удаление, если есть позиции в заказах
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT count(*)::int
                        FROM order_items oi
                        JOIN supplier_items si ON si.id = oi.supplier_item_id
                        WHERE si.supplier_id = %s;
                        """,
                        (supplier_id,),
                    )
                    cnt = _scalar(cur.fetchone())
                    if cnt and cnt > 0:
                        return jsonify(
                            {
                                "error": "supplier has order_items",
                                "details": f"Нельзя удалить: есть позиции в заказах ({cnt}).",
                            }
                        ), 409

                with conn.cursor() as cur:
                    cur.execute("SELECT storage_path FROM supplier_uploads WHERE supplier_id=%s;", (supplier_id,))
                    r = cur.fetchone()
                    if r:
                        stored_path = r[0]

                # чистим зависимости (1 поставщик = 1 прайс)
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM supplier_items WHERE supplier_id=%s;", (supplier_id,))
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM price_list_files WHERE supplier_id=%s;", (supplier_id,))
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM supplier_uploads WHERE supplier_id=%s;", (supplier_id,))
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM suppliers WHERE id=%s;", (supplier_id,))

                conn.commit()

            _safe_remove(stored_path)
            _safe_rmtree(sup_dir)
            return jsonify({"status": "ok", "supplier_id": supplier_id})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    # ---------------- API: list sheets ----------------
    @app.route("/api/sheets", methods=["POST"])
    def api_sheets():
        if "file" not in request.files:
            return jsonify({"error": "file is required"}), 400
        f = request.files["file"]
        if not f or not f.filename:
            return jsonify({"error": "empty filename"}), 400

        raw_filename = os.path.basename(f.filename or "")
        ext = os.path.splitext(raw_filename)[1].lower()
        if ext not in (".xlsx", ".xlsm", ".xls"):
            return jsonify({"error": "unsupported format", "details": "Неподдерживаемый формат. Загрузите .xlsx или .xls"}), 400

        name = f"upload_{uuid4().hex}{ext}"

        tmp_path = os.path.join(UPLOAD_DIR, f"__tmp_sheets__{os.getpid()}_{name}")
        f.save(tmp_path)
        try:
            sheets = import_price.list_excel_sheets(tmp_path)
            return jsonify({"sheets": sheets})
        except Exception as e:
            return jsonify({"error": "failed to read sheets", "details": str(e)}), 400
        finally:
            _safe_remove(tmp_path)

    # ---------------- API: upload (auto / selected sheets) ----------------
    @app.route("/api/upload/<int:supplier_id>", methods=["POST"])
    def api_upload(supplier_id: int):
        if "file" not in request.files:
            return jsonify({"error": "file is required"}), 400
        f = request.files["file"]
        if not f or not f.filename:
            return jsonify({"error": "empty filename"}), 400

        sheet_mode = (request.form.get("sheet_mode") or "all").strip()
        sheets_raw = (request.form.get("sheets") or "").strip()

        # sheets можно передавать CSV строкой или JSON-массивом
        sheet_names: List[str] = []
        if sheets_raw:
            try:
                j = json.loads(sheets_raw)
                if isinstance(j, list):
                    sheet_names = [str(x).strip() for x in j if str(x).strip()]
                else:
                    sheet_names = [s.strip() for s in sheets_raw.split(",") if s.strip()]
            except Exception:
                sheet_names = [s.strip() for s in sheets_raw.split(",") if s.strip()]

        raw_filename = os.path.basename(f.filename or "")
        ext = os.path.splitext(raw_filename)[1].lower()
        allowed_exts = {".xlsx", ".xlsm", ".xls", ".csv"}
        if ext not in allowed_exts:
            return jsonify({"error": "unsupported format", "details": "Неподдерживаемый формат. Загрузите .xlsx или .xls"}), 400

        original = raw_filename or f"upload_{uuid4().hex}{ext}"

        # куда сохраняем
        sup_dir = os.path.join(UPLOAD_DIR, str(supplier_id))
        os.makedirs(sup_dir, exist_ok=True)

        storage_name = f"upload_{uuid4().hex}{ext}"
        dst_path = os.path.join(sup_dir, secure_filename(storage_name))
        base, ext = os.path.splitext(dst_path)
        i = 1
        while os.path.exists(dst_path):
            dst_path = f"{base}({i}){ext}"
            i += 1

        f.save(dst_path)

        old_path = None
        try:
            with db_connect() as conn:
                ensure_schema(conn)
                with conn.cursor() as cur:
                    cur.execute("SELECT storage_path FROM supplier_uploads WHERE supplier_id=%s;", (supplier_id,))
                    r = cur.fetchone()
                    if r:
                        old_path = r[0]

            # импорт в БД
            result = import_price.import_price_file(
                supplier_id=supplier_id,
                file_path=dst_path,
                original_filename=original,
                sheet_mode=sheet_mode,
                sheet_names=sheet_names,
            )

            # записываем мету
            with db_connect() as conn:
                ensure_schema(conn)
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO supplier_uploads (supplier_id, last_uploaded_at, last_filename, storage_path, last_sheet_mode, last_sheets)
                        VALUES (%s, NOW(), %s, %s, %s, %s)
                        ON CONFLICT (supplier_id) DO UPDATE SET
                          last_uploaded_at = EXCLUDED.last_uploaded_at,
                          last_filename = EXCLUDED.last_filename,
                          storage_path = EXCLUDED.storage_path,
                          last_sheet_mode = EXCLUDED.last_sheet_mode,
                          last_sheets = EXCLUDED.last_sheets;
                        """,
                        (
                            supplier_id,
                            original,
                            dst_path,
                            sheet_mode,
                            ",".join(sheet_names) if sheet_names else None,
                        ),
                    )
                conn.commit()

            # удаляем старый файл (по требованию: 1 файл на поставщика)
            if old_path and os.path.abspath(old_path) != os.path.abspath(dst_path):
                _safe_remove(old_path)

            return jsonify(
                {
                    "status": "ok",
                    "supplier_id": supplier_id,
                    "filename": original,
                    **result,
                }
            )
        except Exception as e:
            app.logger.exception("Failed to import upload for supplier %s", supplier_id)
            # если импорт упал — файл оставляем (чтобы можно было скачать/проверить), но можно удалить:
            # _safe_remove(dst_path)
            return jsonify({"error": "upload/import failed", "details": "Ошибка загрузки: не удалось прочитать файл"}), 500

    # ---------------- Search ----------------
    @app.route("/search", methods=["GET"])
    def search():
        q = (request.args.get("q") or "").strip()
        supplier_id = (request.args.get("supplier_id") or "").strip()
        sort = (request.args.get("sort") or "rank").strip()
        limit = request.args.get("limit") or "60"
        category_ids_raw = request.args.get("category_ids") or request.args.get("category_id")

        try:
            limit_i = int(limit)
        except Exception:
            limit_i = 60
        limit_i = max(1, min(limit_i, 300))

        params: Dict[str, Any] = {"limit": limit_i}
        where = ["si.is_active IS TRUE"]

        if category_ids_raw:
            try:
                cids = [int(c) for c in str(category_ids_raw).split(",") if str(c).strip()]
            except Exception:
                cids = []
            if cids:
                params["category_ids"] = tuple(cids)
                where.append("si.category_id = ANY(%(category_ids)s)")

        if supplier_id:
            params["supplier_id"] = int(supplier_id)
            where.append("si.supplier_id = %(supplier_id)s")

        # whitelist сортировки
        if sort not in ("rank", "price_asc", "price_desc", "ppu_asc"):
            sort = "rank"

        if q:
            params["q"] = q
            order_by = {
                "rank": "rank DESC, si.price ASC NULLS LAST, si.id DESC",
                "price_asc": "si.price ASC NULLS LAST, rank DESC, si.id DESC",
                "price_desc": "si.price DESC NULLS LAST, rank DESC, si.id DESC",
                "ppu_asc": "si.price_per_unit ASC NULLS LAST, rank DESC, si.id DESC",
            }[sort]

            sql = f"""
                SELECT
                  si.id,
                  si.supplier_id,
                  s.name AS supplier_name,
                  si.name_raw,
                  si.unit,
                  si.price,
                  si.base_unit,
                  si.base_qty,
                  si.price_per_unit,
                  si.category_id,
                  ts_rank(to_tsvector('russian', si.name_raw),
                          websearch_to_tsquery('russian', %(q)s)) AS rank
                FROM supplier_items si
                JOIN suppliers s ON s.id = si.supplier_id
                WHERE {" AND ".join(where)}
                  AND (
                    to_tsvector('russian', si.name_raw) @@ websearch_to_tsquery('russian', %(q)s)
                    OR si.name_raw ILIKE '%%' || %(q)s || '%%'
                  )
                ORDER BY {order_by}
                LIMIT %(limit)s;
            """
        else:
            order_by = {
                "rank": "si.id DESC",
                "price_asc": "si.price ASC NULLS LAST, si.id DESC",
                "price_desc": "si.price DESC NULLS LAST, si.id DESC",
                "ppu_asc": "si.price_per_unit ASC NULLS LAST, si.id DESC",
            }[sort]

            sql = f"""
                SELECT
                  si.id,
                  si.supplier_id,
                  s.name AS supplier_name,
                  si.name_raw,
                  si.unit,
                  si.price,
                  si.base_unit,
                  si.base_qty,
                  si.price_per_unit,
                  si.category_id,
                  0.0::float AS rank
                FROM supplier_items si
                JOIN suppliers s ON s.id = si.supplier_id
                WHERE {" AND ".join(where)}
                ORDER BY {order_by}
                LIMIT %(limit)s;
            """

        try:
            with db_connect() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(sql, params)
                    rows = cur.fetchall()
            items = [{k: _json_safe(v) for k, v in dict(r).items()} for r in rows]
            return jsonify({"q": q, "items": items})
        except Exception as e:
            return jsonify({"error": "internal error", "details": str(e)}), 500

    # ---------------- Export order to XLSX/CSV ----------------
    @app.route("/export", methods=["POST"])
    def export_order():
        data = request.get_json(silent=True) or {}
        items = data.get("items") or []

        # нормализация входа
        norm: List[Dict[str, Any]] = []
        for it in items:
            if not isinstance(it, dict):
                continue
            supplier = (it.get("supplier_name") or it.get("supplier") or "").strip() or "—"
            name = (it.get("name_raw") or it.get("name") or "").strip()
            if not name:
                continue
            unit = (it.get("unit") or "").strip()
            try:
                qty = float(it.get("qty") or 0) or 0.0
            except Exception:
                qty = 0.0
            try:
                price = float(it.get("price") or 0) or 0.0
            except Exception:
                price = 0.0
            if qty <= 0:
                continue
            norm.append(
                {
                    "supplier_name": supplier,
                    "name_raw": name,
                    "unit": unit,
                    "qty": qty,
                    "price": price,
                }
            )

        if not norm:
            return ("empty", 400)

        norm.sort(key=lambda x: (x["supplier_name"], x["name_raw"]))

        headers = ["Поставщик", "Товар", "Кол-во", "Ед.", "Цена, руб", "Сумма, руб"]

        if Workbook is None:
            # CSV fallback
            import csv
            import io

            out = BytesIO()
            wrapper = io.TextIOWrapper(out, encoding="utf-8", newline="")
            w = csv.writer(wrapper, delimiter=";")
            w.writerow(headers)
            for it in norm:
                s = round(it["price"] * it["qty"], 2)
                w.writerow([it["supplier_name"], it["name_raw"], it["qty"], it["unit"], it["price"], s])
            wrapper.flush()
            out.seek(0)
            return send_file(out, mimetype="text/csv; charset=utf-8", as_attachment=True, download_name="zakaz.csv")

        wb = Workbook()
        ws = wb.active
        ws.title = "Заявка"

        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

        for it in norm:
            s = round(it["price"] * it["qty"], 2)
            ws.append([it["supplier_name"], it["name_raw"], it["qty"], it["unit"], it["price"], s])

        widths = [22, 60, 10, 10, 14, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="zakaz.xlsx",
        )

    # ---------------- 404 ----------------
    @app.errorhandler(404)
    def _not_found(_e):
        p = request.path or ""
        if p.startswith("/api/") or p in ("/search", "/health", "/export"):
            return jsonify({"error": "not found"}), 404
        return render_template("search.html", title=APP_TITLE, active="search"), 200

    app.db_connect = db_connect
    app.run_migrations = run_migrations
    return app


if __name__ == "__main__":
    app = create_app()
    if MIGRATION_CLI or os.getenv("RUN_MIGRATIONS", "").strip().lower() in {"1", "true", "yes", "on"}:
        try:
            with app.db_connect() as conn:
                app.run_migrations(conn)
        except Exception:
            app.logger.exception("Migration run failed")
            sys.exit(1)
        sys.exit(0)
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=False)
else:
    app = create_app()
